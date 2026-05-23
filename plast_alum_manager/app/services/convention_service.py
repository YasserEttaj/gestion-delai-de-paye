from __future__ import annotations

from datetime import date, datetime, timedelta
from pathlib import Path

from sqlalchemy import or_

from app.database.db import session_scope
from app.models.convention_model import Convention
from app.services.log_service import LogService
from config import EXPORT_DIR


CONVENTION_STATUS_ACTIVE = "active"
CONVENTION_STATUS_WARNING = "warning"
CONVENTION_STATUS_EXPIRED = "expired"
CONVENTION_STATUS_COMPLETED = "completed"

CONVENTION_STATUS_LABELS = {
    CONVENTION_STATUS_ACTIVE: "Active",
    CONVENTION_STATUS_WARNING: "Proche échéance",
    CONVENTION_STATUS_EXPIRED: "Expirée",
    CONVENTION_STATUS_COMPLETED: "Terminée",
}

CONVENTION_STATUS_COLORS = {
    CONVENTION_STATUS_ACTIVE: "#22C55E",
    CONVENTION_STATUS_WARNING: "#F59E0B",
    CONVENTION_STATUS_EXPIRED: "#DC2626",
    CONVENTION_STATUS_COMPLETED: "#64748B",
}

DEADLINE_OPTIONS = (60, 90, 120)


class ConventionService:
    @staticmethod
    def parse_date(value: str | date | datetime | None) -> date | None:
        if value in (None, ""):
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        text = str(value).strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
        return None

    @staticmethod
    def calculate(start_date: str | date, deadline_days: int, completed: bool = False, today: date | None = None) -> dict:
        parsed_start = ConventionService.parse_date(start_date)
        if not parsed_start:
            raise ValueError("Date de début invalide.")
        deadline_days = int(deadline_days)
        if deadline_days <= 0:
            raise ValueError("Le délai doit être un nombre de jours positif.")

        due_date = parsed_start + timedelta(days=deadline_days)
        current_date = today or date.today()
        remaining_days = (due_date - current_date).days
        if completed:
            status = CONVENTION_STATUS_COMPLETED
        elif remaining_days > 15:
            status = CONVENTION_STATUS_ACTIVE
        elif remaining_days >= 1:
            status = CONVENTION_STATUS_WARNING
        else:
            status = CONVENTION_STATUS_EXPIRED
        return {
            "start_date": parsed_start.isoformat(),
            "deadline_days": deadline_days,
            "due_date": due_date.isoformat(),
            "remaining_days": remaining_days,
            "status": status,
        }

    @staticmethod
    def _urgency_rank(convention: Convention) -> tuple[int, int, str]:
        if convention.status == CONVENTION_STATUS_EXPIRED:
            return (0, convention.remaining_days, convention.due_date)
        if convention.status == CONVENTION_STATUS_WARNING:
            if convention.remaining_days <= 3:
                return (1, convention.remaining_days, convention.due_date)
            if convention.remaining_days <= 7:
                return (2, convention.remaining_days, convention.due_date)
            return (3, convention.remaining_days, convention.due_date)
        if convention.status == CONVENTION_STATUS_COMPLETED:
            return (5, convention.remaining_days, convention.due_date)
        return (4, convention.remaining_days, convention.due_date)

    @classmethod
    def _apply_calculation(cls, convention: Convention, completed: bool | None = None) -> None:
        is_completed = convention.status == CONVENTION_STATUS_COMPLETED if completed is None else completed
        result = cls.calculate(convention.start_date, convention.deadline_days, completed=is_completed)
        convention.start_date = result["start_date"]
        convention.deadline_days = result["deadline_days"]
        convention.due_date = result["due_date"]
        convention.remaining_days = result["remaining_days"]
        convention.status = result["status"]

    @staticmethod
    def _validate_text(data: dict) -> dict:
        cleaned = {
            "company_name": (data.get("company_name") or "").strip(),
            "convention_number": (data.get("convention_number") or "").strip(),
            "convention_type": (data.get("convention_type") or "").strip(),
            "start_date": (data.get("start_date") or "").strip(),
            "deadline_days": data.get("deadline_days"),
            "notes": (data.get("notes") or "").strip() or None,
        }
        if not cleaned["company_name"] or not cleaned["convention_number"] or not cleaned["convention_type"]:
            raise ValueError("Société, numéro de convention et type sont obligatoires.")
        return cleaned

    @classmethod
    def list_conventions(cls, filters: dict | None = None) -> list[Convention]:
        filters = filters or {}
        cls.recalculate_all()
        with session_scope() as session:
            query = session.query(Convention)
            search = (filters.get("search") or "").strip()
            if search:
                like = f"%{search}%"
                query = query.filter(or_(Convention.company_name.ilike(like), Convention.convention_number.ilike(like)))
            if filters.get("status"):
                query = query.filter(Convention.status == filters["status"])
            if filters.get("deadline_days"):
                query = query.filter(Convention.deadline_days == int(filters["deadline_days"]))
            if filters.get("convention_type"):
                query = query.filter(Convention.convention_type == filters["convention_type"])
            rows = list(query.all())
        rows.sort(key=cls._urgency_rank)
        return rows

    @staticmethod
    def get_convention(convention_id: int) -> Convention | None:
        with session_scope() as session:
            return session.get(Convention, convention_id)

    @staticmethod
    def _check_duplicate(session, convention_number: str, current_id: int | None = None) -> None:
        query = session.query(Convention).filter(Convention.convention_number == convention_number)
        if current_id:
            query = query.filter(Convention.id != current_id)
        if query.first():
            raise ValueError("Ce numéro de convention existe déjà.")

    @classmethod
    def create_convention(cls, data: dict, user_id: int | None) -> Convention:
        cleaned = cls._validate_text(data)
        calculated = cls.calculate(cleaned["start_date"], int(cleaned["deadline_days"]))
        with session_scope() as session:
            cls._check_duplicate(session, cleaned["convention_number"])
            convention = Convention(
                company_name=cleaned["company_name"],
                convention_number=cleaned["convention_number"],
                convention_type=cleaned["convention_type"],
                start_date=calculated["start_date"],
                deadline_days=calculated["deadline_days"],
                due_date=calculated["due_date"],
                remaining_days=calculated["remaining_days"],
                status=calculated["status"],
                notes=cleaned["notes"],
            )
            session.add(convention)
            session.flush()
            LogService.log(session, user_id, "Add convention", "convention", convention.id, convention.convention_number)
            return convention

    @classmethod
    def update_convention(cls, convention_id: int, data: dict, user_id: int | None) -> None:
        cleaned = cls._validate_text(data)
        with session_scope() as session:
            convention = session.get(Convention, convention_id)
            if not convention:
                raise ValueError("Convention introuvable.")
            cls._check_duplicate(session, cleaned["convention_number"], convention_id)
            calculated = cls.calculate(
                cleaned["start_date"],
                int(cleaned["deadline_days"]),
                completed=convention.status == CONVENTION_STATUS_COMPLETED,
            )
            convention.company_name = cleaned["company_name"]
            convention.convention_number = cleaned["convention_number"]
            convention.convention_type = cleaned["convention_type"]
            convention.start_date = calculated["start_date"]
            convention.deadline_days = calculated["deadline_days"]
            convention.due_date = calculated["due_date"]
            convention.remaining_days = calculated["remaining_days"]
            convention.status = calculated["status"]
            convention.notes = cleaned["notes"]
            LogService.log(session, user_id, "Edit convention", "convention", convention.id, convention.convention_number)

    @staticmethod
    def delete_convention(convention_id: int, user_id: int | None) -> None:
        with session_scope() as session:
            convention = session.get(Convention, convention_id)
            if not convention:
                raise ValueError("Convention introuvable.")
            LogService.log(session, user_id, "Delete convention", "convention", convention.id, convention.convention_number)
            session.delete(convention)

    @staticmethod
    def mark_completed(convention_id: int, user_id: int | None) -> None:
        with session_scope() as session:
            convention = session.get(Convention, convention_id)
            if not convention:
                raise ValueError("Convention introuvable.")
            convention.status = CONVENTION_STATUS_COMPLETED
            LogService.log(session, user_id, "Complete convention", "convention", convention.id, convention.convention_number)

    @classmethod
    def recalculate_all(cls, user_id: int | None = None) -> int:
        changed = 0
        with session_scope() as session:
            for convention in session.query(Convention).all():
                old = (convention.due_date, convention.remaining_days, convention.status)
                cls._apply_calculation(convention)
                new = (convention.due_date, convention.remaining_days, convention.status)
                if old != new:
                    changed += 1
            if user_id is not None:
                LogService.log(session, user_id, "Recalculate conventions", "convention", None, f"{changed} mise(s) à jour")
        return changed

    @classmethod
    def stats(cls) -> dict:
        rows = cls.list_conventions({})
        active = [row for row in rows if row.status == CONVENTION_STATUS_ACTIVE]
        warning = [row for row in rows if row.status == CONVENTION_STATUS_WARNING]
        expired = [row for row in rows if row.status == CONVENTION_STATUS_EXPIRED]
        completed = [row for row in rows if row.status == CONVENTION_STATUS_COMPLETED]
        urgent = [row for row in rows if row.status in {CONVENTION_STATUS_WARNING, CONVENTION_STATUS_EXPIRED}]
        return {
            "total": len(rows),
            "active": len(active),
            "warning": len(warning),
            "expired": len(expired),
            "completed": len(completed),
            "urgent": len(urgent),
            "nearest": [row for row in rows if row.status != CONVENTION_STATUS_COMPLETED][:6],
            "expired_rows": expired[:6],
        }

    @classmethod
    def notifications(cls) -> list[dict]:
        notes = []
        for convention in cls.list_conventions({}):
            if convention.status == CONVENTION_STATUS_COMPLETED or convention.remaining_days > 15:
                continue
            if convention.remaining_days <= 0:
                level = "critical"
                label = "Échéance expirée"
            elif convention.remaining_days <= 3:
                level = "critical"
                label = "Échéance ≤ 3 jours"
            elif convention.remaining_days <= 7:
                level = "urgent"
                label = "Échéance ≤ 7 jours"
            else:
                level = "attention"
                label = "Échéance ≤ 15 jours"
            notes.append(
                {
                    "level": level,
                    "title": label,
                    "message": f"{convention.company_name} - {convention.convention_number} - {convention.remaining_days} jour(s)",
                    "convention_id": convention.id,
                }
            )
        return notes[:50]

    @classmethod
    def export_excel(cls, rows: list[Convention], user_id: int | None = None) -> Path:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

        EXPORT_DIR.mkdir(parents=True, exist_ok=True)
        path = EXPORT_DIR / f"conventions_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Conventions"
        headers = ["ID", "Société", "Numéro", "Type", "Début", "Délai", "Échéance", "Jours restants", "Statut", "Notes"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2563EB")
        fill_map = {
            CONVENTION_STATUS_ACTIVE: "DCFCE7",
            CONVENTION_STATUS_WARNING: "FEF3C7",
            CONVENTION_STATUS_EXPIRED: "FEE2E2",
            CONVENTION_STATUS_COMPLETED: "DBEAFE",
        }
        for convention in rows:
            ws.append(
                [
                    convention.id,
                    convention.company_name,
                    convention.convention_number,
                    convention.convention_type,
                    convention.start_date,
                    convention.deadline_days,
                    convention.due_date,
                    convention.remaining_days,
                    CONVENTION_STATUS_LABELS.get(convention.status, convention.status),
                    convention.notes or "",
                ]
            )
            ws.cell(ws.max_row, 9).fill = PatternFill("solid", fgColor=fill_map.get(convention.status, "FFFFFF"))
        for idx, column in enumerate(ws.columns, start=1):
            max_len = max(len(str(cell.value or "")) for cell in column)
            ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 12), 42)
        ws.freeze_panes = "A2"
        wb.save(path)
        with session_scope() as session:
            LogService.log(session, user_id, "Export conventions Excel", "convention", None, str(path))
        return path

    @classmethod
    def export_pdf(cls, rows: list[Convention], user_name: str, user_id: int | None = None) -> Path:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

        EXPORT_DIR.mkdir(parents=True, exist_ok=True)
        path = EXPORT_DIR / f"conventions_{datetime.now():%Y%m%d_%H%M%S}.pdf"
        doc = SimpleDocTemplate(str(path), pagesize=landscape(A4), rightMargin=24, leftMargin=24, topMargin=24, bottomMargin=24)
        styles = getSampleStyleSheet()
        story = [
            Paragraph("<b>PLAST ALUM</b>", styles["Title"]),
            Paragraph("Suivi des conventions et échéances configurables", styles["Heading2"]),
            Paragraph(f"Exporté le {datetime.now():%d/%m/%Y %H:%M} par {user_name}", styles["Normal"]),
            Spacer(1, 12),
        ]
        data = [["Société", "Numéro", "Type", "Début", "Délai", "Échéance", "Jours", "Statut"]]
        for convention in rows:
            data.append(
                [
                    convention.company_name[:28],
                    convention.convention_number,
                    convention.convention_type[:22],
                    convention.start_date,
                    str(convention.deadline_days),
                    convention.due_date,
                    str(convention.remaining_days),
                    CONVENTION_STATUS_LABELS.get(convention.status, convention.status),
                ]
            )
        if len(data) == 1:
            data.append(["Aucune donnée", "", "", "", "", "", "", ""])
        table = Table(data, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ]
            )
        )
        story.append(table)
        story.append(Spacer(1, 12))
        story.append(Paragraph("Note: les délais sont configurables et ne constituent pas un avis juridique ou fiscal.", styles["Normal"]))
        doc.build(story)
        with session_scope() as session:
            LogService.log(session, user_id, "Export conventions PDF", "convention", None, str(path))
        return path
