from __future__ import annotations

from datetime import datetime
from pathlib import Path
import unicodedata

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.database.db import session_scope
from app.models.invoice_model import Invoice
from app.models.supplier_model import Supplier
from app.models.payment_model import Payment
from app.services.deadline_service import DeadlineService
from app.services.invoice_service import InvoiceService
from app.services.log_service import LogService
from config import EXPORT_DIR, STATUS_LABELS_FR, STATUS_PAID, STATUS_PARTIAL, STATUS_UNPAID


class ExcelService:
    EXPECTED_COLUMNS = [
        "Fournisseur",
        "Numéro facture",
        "Date facture",
        "Date réception",
        "Montant HT",
        "TVA",
        "Montant TTC",
        "Statut",
        "Date paiement",
        "Mode paiement",
        "Notes",
    ]

    STATUS_IMPORT = {
        "payée": STATUS_PAID,
        "payee": STATUS_PAID,
        "paid": STATUS_PAID,
        "non payée": STATUS_UNPAID,
        "non payee": STATUS_UNPAID,
        "unpaid": STATUS_UNPAID,
        "partiellement payée": STATUS_PARTIAL,
        "partiellement payee": STATUS_PARTIAL,
        "partial": STATUS_PARTIAL,
    }

    @staticmethod
    def _normalize_status(value: object) -> str:
        text = str(value or "Non payée").strip().lower()
        text = unicodedata.normalize("NFKD", text)
        text = "".join(char for char in text if not unicodedata.combining(char))
        return text.replace("?", "e")

    @staticmethod
    def _clean_float(value, default: float = 0.0) -> float:
        if value in (None, ""):
            return default
        try:
            return float(value)
        except (TypeError, ValueError):
            return default

    @staticmethod
    def _clean_date(value) -> str | None:
        if value in (None, ""):
            return None
        if hasattr(value, "date"):
            return value.date().isoformat()
        text = str(value).strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(text, fmt).date().isoformat()
            except ValueError:
                continue
        parsed = DeadlineService.parse_date(text)
        return parsed.isoformat() if parsed else None

    @staticmethod
    def export_invoices(rows: list[dict], filename: str | None = None, user_id: int | None = None) -> Path:
        EXPORT_DIR.mkdir(parents=True, exist_ok=True)
        filename = filename or f"rapport_factures_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        path = EXPORT_DIR / filename
        wb = Workbook()
        ws = wb.active
        ws.title = "Factures"
        headers = [
            "ID",
            "Fournisseur",
            "Numéro facture",
            "Date facture",
            "Date réception",
            "Montant TTC",
            "Statut",
            "Date paiement",
            "Jours",
            "Catégorie délai",
            "Montant payé",
            "Reste à payer",
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2563EB")
        fill_map = {
            "normal": "DCFCE7",
            "attention": "FEF3C7",
            "urgent": "FEE2E2",
            "critical": "7F1D1D",
        }
        for row in rows:
            invoice = row["invoice"]
            deadline = row["deadline"]
            ws.append(
                [
                    invoice.id,
                    row["supplier"].name,
                    invoice.invoice_number,
                    invoice.invoice_date,
                    invoice.reception_date,
                    float(invoice.amount_ttc or 0),
                    STATUS_LABELS_FR.get(invoice.status, invoice.status),
                    invoice.payment_date,
                    deadline.days,
                    deadline.label,
                    row["paid_amount"],
                    row["outstanding_amount"],
                ]
            )
            deadline_cell = ws.cell(ws.max_row, 10)
            deadline_cell.fill = PatternFill("solid", fgColor=fill_map.get(deadline.category, "FFFFFF"))
            if deadline.category == "critical":
                deadline_cell.font = Font(color="FFFFFF", bold=True)
        total_row = ws.max_row + 1
        ws.cell(total_row, 5, "Totaux")
        ws.cell(total_row, 6, f"=SUM(F2:F{total_row-1})")
        ws.cell(total_row, 11, f"=SUM(K2:K{total_row-1})")
        ws.cell(total_row, 12, f"=SUM(L2:L{total_row-1})")
        for cell in ws[total_row]:
            cell.font = Font(bold=True)
        for idx, column in enumerate(ws.columns, start=1):
            max_len = max(len(str(cell.value or "")) for cell in column)
            ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 12), 40)
        ws.freeze_panes = "A2"
        wb.save(path)
        with session_scope() as session:
            LogService.log(session, user_id, "Export Excel", "report", None, str(path))
        return path

    @classmethod
    def preview_import(cls, file_path: str) -> dict:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
        rows = []
        errors = []
        for expected in cls.EXPECTED_COLUMNS[:7]:
            if expected not in headers:
                errors.append(f"Colonne manquante: {expected}")
        header_index = {name: idx for idx, name in enumerate(headers)}
        with session_scope() as session:
            suppliers = {supplier.name.lower(): supplier for supplier in session.query(Supplier).all()}
            duplicates = {
                (invoice.supplier_id, invoice.invoice_number.lower())
                for invoice in session.query(Invoice).all()
            }
            file_invoice_keys: set[tuple[str, str]] = set()
            for excel_row in range(2, ws.max_row + 1):
                raw = {header: ws.cell(excel_row, idx + 1).value for header, idx in header_index.items()}
                row_errors = []
                supplier_name = str(raw.get("Fournisseur") or "").strip()
                invoice_number = str(raw.get("Numéro facture") or "").strip()
                invoice_date = cls._clean_date(raw.get("Date facture"))
                amount_ttc = raw.get("Montant TTC")
                amount_ht = raw.get("Montant HT")
                if not supplier_name:
                    row_errors.append("Fournisseur obligatoire")
                if not invoice_number:
                    row_errors.append("Numéro facture obligatoire")
                if not invoice_date:
                    row_errors.append("Date facture obligatoire")
                try:
                    amount_ttc = float(amount_ttc if amount_ttc not in (None, "") else amount_ht)
                    if amount_ttc <= 0:
                        row_errors.append("Montant invalide")
                except (TypeError, ValueError):
                    row_errors.append("Montant invalide")
                    amount_ttc = 0.0
                supplier = suppliers.get(supplier_name.lower())
                if supplier and (supplier.id, invoice_number.lower()) in duplicates:
                    row_errors.append("Doublon déjà existant")
                file_key = (supplier_name.lower(), invoice_number.lower())
                if supplier_name and invoice_number:
                    if file_key in file_invoice_keys:
                        row_errors.append("Doublon dans le fichier")
                    file_invoice_keys.add(file_key)
                status_text = cls._normalize_status(raw.get("Statut"))
                status = cls.STATUS_IMPORT.get(status_text, STATUS_UNPAID)
                payment_date = cls._clean_date(raw.get("Date paiement"))
                payment_method = str(raw.get("Mode paiement") or "").strip()
                if status == STATUS_PAID and not payment_date:
                    row_errors.append("Date paiement obligatoire pour une facture payée")
                if status == STATUS_PAID and not payment_method:
                    row_errors.append("Mode paiement obligatoire pour une facture payée")
                amount_ht_value = cls._clean_float(raw.get("Montant HT"))
                tva_rate_value = cls._clean_float(raw.get("TVA"), 20.0)
                rows.append(
                    {
                        "line": excel_row,
                        "raw": raw,
                        "supplier_name": supplier_name,
                        "invoice_number": invoice_number,
                        "invoice_date": invoice_date,
                        "reception_date": cls._clean_date(raw.get("Date réception")),
                        "amount_ht": amount_ht_value,
                        "tva_rate": tva_rate_value,
                        "amount_ttc": float(amount_ttc),
                        "status": status,
                        "payment_date": payment_date,
                        "payment_method": payment_method,
                        "notes": str(raw.get("Notes") or ""),
                        "errors": row_errors,
                    }
                )
        return {"headers": headers, "rows": rows, "errors": errors}

    @classmethod
    def import_valid_rows(cls, preview: dict, user_id: int | None, create_suppliers: bool = True) -> dict:
        if preview.get("errors"):
            raise ValueError("Le fichier Excel ne contient pas toutes les colonnes obligatoires.")
        imported = 0
        skipped = 0
        with session_scope() as session:
            suppliers = {supplier.name.lower(): supplier for supplier in session.query(Supplier).all()}
            imported_keys: set[tuple[int, str]] = set()
            for row in preview.get("rows", []):
                if row["errors"]:
                    skipped += 1
                    continue
                supplier = suppliers.get(row["supplier_name"].lower())
                if not supplier and create_suppliers:
                    supplier = Supplier(name=row["supplier_name"])
                    session.add(supplier)
                    session.flush()
                    suppliers[row["supplier_name"].lower()] = supplier
                if not supplier:
                    skipped += 1
                    continue
                duplicate = session.query(Invoice).filter(
                    Invoice.supplier_id == supplier.id,
                    Invoice.invoice_number == row["invoice_number"],
                ).first()
                import_key = (supplier.id, row["invoice_number"].lower())
                if duplicate or import_key in imported_keys:
                    skipped += 1
                    continue
                amount_ht = row["amount_ht"] or round(row["amount_ttc"] / (1 + row["tva_rate"] / 100), 2)
                amount_tva = round(row["amount_ttc"] - amount_ht, 2)
                invoice = Invoice(
                    supplier_id=supplier.id,
                    invoice_number=row["invoice_number"],
                    invoice_date=row["invoice_date"],
                    reception_date=row["reception_date"],
                    due_date=InvoiceService.default_due_date(row["invoice_date"], row["reception_date"]),
                    amount_ht=amount_ht,
                    tva_rate=row["tva_rate"],
                    amount_tva=amount_tva,
                    amount_ttc=row["amount_ttc"],
                    status=row["status"],
                    payment_date=row["payment_date"],
                    payment_method=row["payment_method"],
                    notes=row["notes"],
                    created_by=user_id,
                )
                session.add(invoice)
                session.flush()
                if row["status"] == STATUS_PAID:
                    session.add(
                        Payment(
                            invoice_id=invoice.id,
                            amount=row["amount_ttc"],
                            payment_date=row["payment_date"],
                            payment_method=row["payment_method"],
                            reference="Import Excel",
                            notes=row["notes"],
                        )
                    )
                imported_keys.add(import_key)
                imported += 1
            LogService.log(session, user_id, "Import Excel", "invoice", None, f"{imported} importées, {skipped} ignorées")
        return {"imported": imported, "skipped": skipped}
